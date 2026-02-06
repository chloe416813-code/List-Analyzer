import streamlit as st
import pandas as pd
import io
import zipfile
from datetime import datetime
import sys

# ================= 0. 系統環境檢查 =================
try:
    import openpyxl
    import msoffcrypto
    import xlsxwriter
except ImportError:
    st.error("🛑 缺少必要套件，請檢查 requirements.txt")
    st.stop()

# ================= 1. 核心邏輯區 =================
REF_DATE = datetime(2025, 10, 20)

def parse_roc_birthday(roc_val):
    """ 解析民國年生日 """
    if pd.isna(roc_val): return None
    s = str(roc_val).strip().replace('\t', '').replace(' ', '')
    if s == '' or s.lower() == 'nan': return None
    s_clean = s.replace('年', '.').replace('月', '.').replace('日', '').replace('-', '.').replace('/', '.')
    
    parts = []
    if '.' in s_clean: parts = s_clean.split('.')
    elif s_clean.isdigit():
        if len(s_clean) == 6: parts = [s_clean[:2], s_clean[2:4], s_clean[4:]]
        elif len(s_clean) == 7: parts = [s_clean[:3], s_clean[3:5], s_clean[5:]]
    try:
        if len(parts) != 3: return None
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        if not (1 <= m <= 12 and 1 <= d <= 31): return None
        return datetime(y + 1911, m, d)
    except:
        return None

def calculate_age(born):
    if born is None: return -1
    return REF_DATE.year - born.year - ((REF_DATE.month, REF_DATE.day) < (born.month, born.day))

def open_excel_safe(file_content, password):
    """ 安全開啟 Excel (支援加密) """
    file_stream = io.BytesIO(file_content)
    try:
        return openpyxl.load_workbook(file_stream)
    except:
        file_stream.seek(0)
    
    if password:
        try:
            decrypted = io.BytesIO()
            office_file = msoffcrypto.OfficeFile(file_stream)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return openpyxl.load_workbook(decrypted)
        except:
            return None
    return None

def find_col_name(columns, keywords):
    """ 智慧搜尋欄位名稱 """
    for col in columns:
        col_str = str(col)
        if any(k in col_str for k in keywords):
            return col
    return None

def process_file_logic(filename, content, password):
    """ 
    核心處理：
    1. 檢查並標記黃底 (回傳 Excel binary)
    2. 萃取數據 (回傳 DataFrame 供統計用)
    """
    wb = open_excel_safe(content, password)
    if wb is None:
        return None, None, {"filename": filename, "status": "Fail", "msg": "無法開啟 (密碼錯誤或格式不支援)"}

    ws = wb.active
    
    # 1. 尋找表頭
    header_row_idx = 0
    
    # 掃描前 5 行找關鍵字
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        row_str = [str(c) if c else '' for c in row]
        if any('身分證' in s for s in row_str) or any('生日' in s for s in row_str):
            header_row_idx = r_idx
            break
    
    # 2. 準備統計用的 DataFrame (重新讀取資料)
    data = list(ws.values)
    if not data: return None, None, {"filename": filename, "status": "Fail", "msg": "空檔案"}
    
    header = data[header_row_idx]
    rows = data[header_row_idx+1:]
    df = pd.DataFrame(rows, columns=header)
    
    # 3. 欄位對應
    cols = [str(c) for c in df.columns]
    
    # 定義關鍵字
    key_id = ['身分證', 'ID', '證號']
    key_birth = ['生日', '出生', 'Birth']
    key_city = ['縣市', '城市', 'City', '地區', '居住地']
    key_school = ['學校', '校名', 'School', '單位', '就讀學校']
    key_role = ['職稱', '身分', '身份', 'Role', '職務', '對象']

    col_id = find_col_name(cols, key_id)
    col_birth = find_col_name(cols, key_birth)
    col_city = find_col_name(cols, key_city)
    col_school = find_col_name(cols, key_school)
    col_role = find_col_name(cols, key_role)

    stats_meta = {"filename": filename, "under_15": 0, "adult": 0, "errors": 0, "status": "Success", "msg": "OK"}

    if not col_id or not col_birth:
        return None, None, {"filename": filename, "status": "Fail", "msg": "找不到關鍵欄位 (身分證/生日)"}

    # 4. 檢查邏輯 (使用 openpyxl 標記黃底)
    wb_out = open_excel_safe(content, password)
    ws_out = wb_out.active
    from openpyxl.styles import PatternFill
    YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 找出 openpyxl 中的欄位 index
    op_header = list(ws_out.iter_rows(min_row=header_row_idx+1, max_row=header_row_idx+1, values_only=True))[0]
    op_header = [str(c) for c in op_header]
    
    try:
        idx_id = op_header.index(col_id)
        idx_birth = op_header.index(col_birth)
    except:
        return None, None, {"filename": filename, "status": "Fail", "msg": "欄位索引對應失敗"}

    for row in ws_out.iter_rows(min_row=header_row_idx+2):
        # 生日檢查
        if idx_birth < len(row):
            cell = row[idx_birth]
            dt = parse_roc_birthday(cell.value)
            if dt is None:
                cell.fill = YELLOW
                stats_meta["errors"] += 1
            else:
                age = calculate_age(dt)
                if 0 <= age < 15: stats_meta["under_15"] += 1
                elif age >= 15: stats_meta["adult"] += 1
        
        # 身分證檢查
        if idx_id < len(row):
            cell = row[idx_id]
            val = str(cell.value).strip() if cell.value else ""
            if not val or val == 'None' or len(val) != 10:
                cell.fill = YELLOW
                stats_meta["errors"] += 1

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    
    # 5. 整理統計資料回傳
    df_stat = df.copy()
    
    # 統一欄位名稱
    rename_map = {}
    if col_city: rename_map[col_city] = '縣市'
    else: df_stat['縣市'] = '未填縣市'
    
    if col_school: rename_map[col_school] = '學校'
    else: df_stat['學校'] = '未填學校'
    
    if col_role: rename_map[col_role] = '職稱'
    else: df_stat['職稱'] = '一般'
    
    df_stat.rename(columns=rename_map, inplace=True)
    
    return output, df_stat, stats_meta

# ================= 2. 執行函式 =================

def run_checker_and_stats(files, pwd):
    processed_files = []
    report_list = []
    combined_df = pd.DataFrame() 
    
    bar = st.progress(0)
    for i, f in enumerate(files):
        excel_data, df_stat, meta = process_file_logic(f.name, f.read(), pwd)
        
        report_list.append(meta)
        if excel_data:
            processed_files.append((f"已檢查_{f.name}", excel_data.getvalue()))
        if df_stat is not None:
            df_stat['來源檔案'] = f.name
            combined_df = pd.concat([combined_df, df_stat], ignore_index=True)
            
        bar.progress((i + 1) / len(files))
    return processed_files, report_list, combined_df

def run_encryptor_native(files, pwd):
    """ 強制使用 xlsxwriter 原生寫入，避開 pandas 引擎衝突 """
    processed = []
    bar = st.progress(0)
    for i, f in enumerate(files):
        try:
            content = f.read()
            try:
                df = pd.read_excel(io.BytesIO(content))
            except:
                st.error(f"❌ {f.name}: 讀取失敗。")
                continue
            
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet()
            
            # 寫入
            header = df.columns.values
            for c, val in enumerate(header):
                worksheet.write(0, c, str(val))
            data = df.fillna("").values
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    worksheet.write(r + 1, c, val)
            
            # 加密
            workbook.set_encryption(pwd)
            workbook.close()
            output.seek(0)
            processed.append((f"加密_{f.name}", output.getvalue()))
            
        except Exception as e:
            st.error(f"❌ {f.name} 失敗: {e}")
        bar.progress((i + 1) / len(files))
    return processed

# ================= 3. 主介面 =================

st.set_page_config(page_title="科普列車統計系統 V6.0", page_icon="📊", layout="wide")
st.title("📊 科普列車 - 檢查與統計系統 V6.0")

tab1, tab2 = st.tabs(["🔍 檢查與統計", "🔒 批次加密"])

with tab1:
    st.header("1. 名單檢查 & 自動統計")
    st.info("系統會檢查格式錯誤(黃底)，並自動產出 **Excel 統計報表**。")
    
    c1, c2 = st.columns([1, 2])
    with c1:
        pwd = st.text_input("檔案密碼 (若無則留空)", type="password", key="p1")
    with c2:
        files1 = st.file_uploader("上傳 Excel (支援多檔合併)", type=['xlsx'], accept_multiple_files=True, key="u1")
    
    if files1 and st.button("🚀 開始分析", key="b1"):
        res, rep, big_df = run_checker_and_stats(files1, pwd)
        
        # --- 儀表板 ---
        if not big_df.empty:
            st.divider()
            st.subheader("📈 數據儀表板")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("總參與人數", f"{len(big_df)} 人")
            col2.metric("涵蓋縣市", f"{big_df['縣市'].nunique()} 個")
            col3.metric("參與學校", f"{big_df['學校'].nunique()} 所")
            
            chart1, chart2 = st.columns(2)
            with chart1:
                st.markdown("**各縣市人數**")
                st.bar_chart(big_df['縣市'].value_counts())
            with chart2:
                st.markdown("**職稱/身分比例**")
                st.bar_chart(big_df['職稱'].value_counts(), color="#ffaa00")
            
            # --- 產生統計 Excel ---
            st.divider()
            st.subheader("📥 下載專區")
            
            try:
                pivot = big_df.pivot_table(index=['縣市', '學校'], columns='職稱', aggfunc='size', fill_value=0)
                pivot['該校總計'] = pivot.sum(axis=1)
                
                stats_io = io.BytesIO()
                with pd.ExcelWriter(stats_io, engine='xlsxwriter') as writer:
                    pivot.to_excel(writer, sheet_name='各校統計')
                    big_df['縣市'].value_counts().to_frame(name="人數").to_excel(writer, sheet_name='縣市統計')
                    big_df.to_excel(writer, sheet_name='總名單明細', index=False)
                    writer.sheets['各校統計'].set_column(0, 0, 15)
                    writer.sheets['各校統計'].set_column(1, 1, 30)
                
                stats_io.seek(0)
                
                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button("📊 下載統計報表 (.xlsx)", stats_io, "統計報表.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
                    st.caption("包含：各校詳細統計、縣市分佈、總名單。")
                    
                with col_dl2:
                    if res:
                        z = io.BytesIO()
                        with zipfile.ZipFile(z, "w") as zf:
                            for n, d in res: zf.writestr(n, d)
                            txt = "\n".join([f"{r['filename']}: {r['msg']}" for r in rep])
                            zf.writestr("report.txt", txt)
                        st.download_button("📦 下載已檢查原始檔 (ZIP)", z.getvalue(), "檢查結果_黃底.zip", "application/zip")

            except Exception as e:
                st.error(f"統計報表產生失敗: {e}")

with tab2:
    st.header("2. 批次加密")
    st.warning("請上傳無密碼檔案。")
    new_pwd = st.text_input("設定新密碼", type="password", key="p2")
    files2 = st.file_uploader("上傳要加密的檔案", type=['xlsx'], accept_multiple_files=True, key="u2")
    
    if files2 and new_pwd:
        if st.button("🔒 開始加密", key="b2"):
            res = run_encryptor_native(files2, new_pwd)
            if res:
                st.success(f"處理完成 {len(res)} 個檔案")
                z = io.BytesIO()
                with zipfile.ZipFile(z, "w") as zf:
                    for n, d in res: zf.writestr(n, d)
                st.download_button("📦 下載加密檔案", z.getvalue(), "已加密.zip", "application/zip")
